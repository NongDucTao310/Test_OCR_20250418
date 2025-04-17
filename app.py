import os
import cv2
import tkinter as tk
from tkinter import filedialog, messagebox
from ultralytics import YOLO
import easyocr
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from PIL import Image
import numpy as np
import uuid
import re
import shutil
import threading
import sys
from tkinter.scrolledtext import ScrolledText

class ConsoleRedirector:
    """Redirect console output to a TextBox."""
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)  # Cuộn xuống cuối TextBox

    def flush(self):
        pass  # Không cần thực hiện gì khi flush

class TaskController:
    """Quản lý trạng thái của các tiến trình."""
    def __init__(self):
        self.running = False  # Cờ để kiểm tra tiến trình có đang chạy không

    def start(self):
        self.running = True

    def stop(self):
        self.running = False

    def is_running(self):
        return self.running

def run_task_in_thread(task_function, controller):
    """Chạy một tác vụ trong luồng riêng biệt."""
    def wrapper():
        controller.start()
        try:
            task_function()
        finally:
            controller.stop()
    thread = threading.Thread(target=wrapper)
    thread.daemon = True  # Đảm bảo luồng dừng khi ứng dụng chính dừng
    thread.start()

# Function: Convert to Black and White
def convert_to_bw(image_path, lower_hsv1, upper_hsv1, lower_hsv2, upper_hsv2, kernel_size=(2, 2), threshold=20):
    image = cv2.imread(image_path)
    if image is None:
        print(f"Error: Unable to load image {image_path}.")
        return None

    hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    mask1 = cv2.inRange(hsv_image, np.array(lower_hsv1), np.array(upper_hsv1))
    mask2 = cv2.inRange(hsv_image, np.array(lower_hsv2), np.array(upper_hsv2))
    combined_mask = cv2.bitwise_or(mask1, mask2)
    result_image = cv2.bitwise_and(image, image, mask=combined_mask)
    gray_image = cv2.cvtColor(result_image, cv2.COLOR_RGB2GRAY)
    _, processed_image = cv2.threshold(gray_image, threshold, 255, cv2.THRESH_BINARY)
    kernel1 = cv2.getStructuringElement(cv2.MORPH_RECT, kernel_size)
    processed_image = cv2.morphologyEx(processed_image, cv2.MORPH_OPEN, kernel1)
    kernel2 = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
    processed_image = cv2.morphologyEx(processed_image, cv2.MORPH_CLOSE, kernel2)
    output_image = cv2.cvtColor(processed_image, cv2.COLOR_GRAY2BGR)
    return output_image

def process_folder_bw(input_folder, output_folder, lower_hsv1, upper_hsv1, lower_hsv2, upper_hsv2, kernel_size=(1, 1), threshold=20):
    for root, _, files in os.walk(input_folder):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                input_path = os.path.join(root, file)
                relative_path = os.path.relpath(input_path, input_folder)
                output_path = os.path.join(output_folder, relative_path)
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                processed_image = convert_to_bw(input_path, lower_hsv1, upper_hsv1, lower_hsv2, upper_hsv2, kernel_size, threshold)
                if processed_image is not None:
                    cv2.imwrite(output_path, processed_image)
                    print(f"Processed and saved: {output_path}")

# Function: YOLO Predict
def detect_and_crop(model_path, input_folder, output_folder, conf_threshold=0.25, task_controller=None):
    model = YOLO(model_path)
    image_files = []
    for root, _, files in os.walk(input_folder):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                image_files.append(os.path.join(root, file))
    total_files = len(image_files)
    print(f"Total images to process: {total_files}")

    for idx, input_path in enumerate(image_files, start=1):
        if task_controller and not task_controller.is_running():
            print("Detect and Crop stopped.")
            break
        relative_path = os.path.relpath(input_path, input_folder)
        output_dir = os.path.join(output_folder, os.path.splitext(relative_path)[0])
        os.makedirs(output_dir, exist_ok=True)
        image = cv2.imread(input_path)
        if image is None:
            print(f"Error: Unable to load image {input_path}. Skipping...")
            continue
        results = model.predict(source=input_path, conf=conf_threshold)
        for idx_box, box in enumerate(results[0].boxes.xyxy, start=1):
            if task_controller and not task_controller.is_running():
                print("Detect and Crop stopped.")
                break
            x1, y1, x2, y2 = map(int, box)
            cropped_image = image[y1:y2, x1:x2]
            if cropped_image.size == 0:
                print(f"Error: Cropped image is empty for {input_path}. Skipping...")
                continue
            output_file = os.path.join(output_dir, f"result_{idx_box}.jpg")
            cv2.imwrite(output_file, cropped_image)
            print(f"Saved: {output_file}")
        print(f"Progress: {idx}/{total_files} ({(idx / total_files) * 100:.2f}%)")

# Function: Export to Excel
def clean_text(text):
    match = re.findall(r'-?\d+(?:\.\d+)?', text)
    return " ".join(match)

def read_text_from_images(input_folder, output_excel, task_controller=None):
    reader = easyocr.Reader(['en'], gpu=False)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "OCR Results"
    sheet.append(["Path", "Results", "Text", "Image"])
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 20

    temp_dir = "./temporary"
    os.makedirs(temp_dir, exist_ok=True)
    temp_image_paths = []
    image_files = []
    for root, _, files in os.walk(input_folder):
        for file in files:
            if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                image_files.append(os.path.join(root, file))
    total_files = len(image_files)
    print(f"Total images to process: {total_files}")

    for idx, input_path in enumerate(image_files, start=1):
        if task_controller and not task_controller.is_running():
            print("Read and Export stopped.")
            break
        relative_path = os.path.relpath(input_path, input_folder)
        path_without_filename = os.path.dirname(relative_path)
        result_name = os.path.splitext(os.path.basename(input_path))[0]
        result_text = "No text detected"
        try:
            results = reader.readtext(input_path)
            if results:
                result_text = "\n".join([clean_text(res[1]) for res in results])
        except Exception as e:
            result_text = f"Error: {str(e)}"
        print(f"Image: {relative_path}")
        print(f"Detected Text: {result_text}")
        print(f"Progress: {idx}/{total_files} ({(idx / total_files) * 100:.2f}%)")
        print("-" * 50)
        row = [path_without_filename, result_name, result_text]
        sheet.append(row)
        try:
            img = Image.open(input_path)
            img.thumbnail((100, 100))
            temp_image_path = os.path.join(temp_dir, f"temp_thumbnail_{uuid.uuid4().hex}.png")
            img.save(temp_image_path)
            temp_image_paths.append(temp_image_path)
            excel_img = ExcelImage(temp_image_path)
            excel_img.anchor = f"D{sheet.max_row}"
            sheet.add_image(excel_img)
        except Exception as e:
            print(f"Error adding image {input_path} to Excel: {str(e)}")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        sheet.row_dimensions[row[0].row].height = 30

    workbook.save(output_excel)
    print(f"Results saved to {output_excel}")
    for temp_image_path in temp_image_paths:
        if os.path.exists(temp_image_path):
            os.remove(temp_image_path)
    if os.path.exists(temp_dir) and not os.listdir(temp_dir):
        os.rmdir(temp_dir)

# Main Application
def run_app():
    root = tk.Tk()
    root.title("Image Processing Application")

    # Variables for user inputs
    input_folder_var = tk.StringVar()
    model_path_var = tk.StringVar(value="./model/best.pt")  # Mặc định là ./model/best.pt
    output_excel_var = tk.StringVar(value="./ocr_results.xlsx")
    conf_threshold_var = tk.DoubleVar(value=0.6)

    # Variables for adjustable parameters
    lower_hsv1_var = tk.StringVar(value="0,180,100")
    upper_hsv1_var = tk.StringVar(value="5,255,255")
    lower_hsv2_var = tk.StringVar(value="175,180,100")
    upper_hsv2_var = tk.StringVar(value="180,255,255")
    kernel_size_var = tk.StringVar(value="1,1")
    threshold_var = tk.IntVar(value=20)

    # Task controller
    task_controller = TaskController()

    def select_input_folder():
        folder = filedialog.askdirectory(title="Select Input Folder")
        if folder:
            input_folder_var.set(folder)

    def select_model_file():
        file = filedialog.askopenfilename(title="Select Model File", filetypes=[("YOLO Model", "*.pt")])
        if file:
            model_path_var.set(file)

    def parse_tuple(value):
        """Parse a string like '1,2,3' into a tuple of integers."""
        return tuple(map(int, value.split(',')))

    def clear_folder(folder):
        """Xóa nội dung trong một thư mục cụ thể."""
        if os.path.exists(folder):
            try:
                shutil.rmtree(folder)  # Xóa toàn bộ thư mục và nội dung bên trong
                print(f"Cleared folder: {folder}")
            except Exception as e:
                print(f"Error clearing folder {folder}: {e}")

    def run_convert_to_bw():
        def task():
            clear_folder("./converted_Img")  # Chỉ xóa nội dung trong ./converted_Img
            input_folder = input_folder_var.get()
            if not input_folder:
                messagebox.showerror("Error", "Please select an input folder!")
                return
            output_bw_folder = "./converted_Img"
            os.makedirs(output_bw_folder, exist_ok=True)
            print("Running Convert to BW...")

            # Parse parameters
            lower_hsv1 = parse_tuple(lower_hsv1_var.get())
            upper_hsv1 = parse_tuple(upper_hsv1_var.get())
            lower_hsv2 = parse_tuple(lower_hsv2_var.get())
            upper_hsv2 = parse_tuple(upper_hsv2_var.get())
            kernel_size = parse_tuple(kernel_size_var.get())
            threshold = threshold_var.get()

            for root, _, files in os.walk(input_folder):
                if not task_controller.is_running():
                    print("Convert to BW stopped.")
                    break
                for file in files:
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                        input_path = os.path.join(root, file)
                        relative_path = os.path.relpath(input_path, input_folder)
                        output_path = os.path.join(output_bw_folder, relative_path)
                        os.makedirs(os.path.dirname(output_path), exist_ok=True)
                        processed_image = convert_to_bw(input_path, lower_hsv1, upper_hsv1, lower_hsv2, upper_hsv2, kernel_size, threshold)
                        if processed_image is not None:
                            cv2.imwrite(output_path, processed_image)
                            print(f"Processed and saved: {output_path}")
            print("Convert to BW completed!")
        run_task_in_thread(task, task_controller)

    def run_detect_and_crop():
        def task():
            clear_folder("./predict_output")
            input_folder = "./converted_Img"
            model_path = model_path_var.get()
            if not model_path:
                messagebox.showerror("Error", "Please select a YOLO model file!")
                return
            output_yolo_folder = "./predict_output"
            os.makedirs(output_yolo_folder, exist_ok=True)
            print("Running YOLO Predict...")
            detect_and_crop(model_path, input_folder, output_yolo_folder, conf_threshold=conf_threshold_var.get(), task_controller=task_controller)
            print("Detect and Crop completed!")
        run_task_in_thread(task, task_controller)

    def run_read_and_export():
        def task():
            clear_folder("./temporary")
            input_folder = "./predict_output"
            output_excel = output_excel_var.get()
            print("Running Export to Excel...")
            read_text_from_images(input_folder, output_excel, task_controller=task_controller)
            print("Read and Export completed!")
        run_task_in_thread(task, task_controller)

    def run_all_tasks():
        def task():
            """Chạy tất cả các tác vụ theo thứ tự."""
            clear_folder("./converted_Img")
            clear_folder("./predict_output")
            clear_folder("./temporary")

            # Chạy Convert to BW
            if task_controller.is_running():
                print("Running Convert to BW...")
                for root, _, files in os.walk(input_folder_var.get()):
                    if not task_controller.is_running():
                        print("Convert to BW stopped.")
                        break
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff')):
                            input_path = os.path.join(root, file)
                            relative_path = os.path.relpath(input_path, input_folder_var.get())
                            output_path = os.path.join("./converted_Img", relative_path)
                            os.makedirs(os.path.dirname(output_path), exist_ok=True)
                            processed_image = convert_to_bw(input_path, (0, 180, 100), (5, 255, 255), (175, 180, 100), (180, 255, 255))
                            if processed_image is not None:
                                cv2.imwrite(output_path, processed_image)
                                print(f"Processed and saved: {output_path}")
                print("Convert to BW completed!")

            # Chạy Detect and Crop
            if task_controller.is_running():
                print("Running YOLO Predict...")
                detect_and_crop(model_path_var.get(), "./converted_Img", "./predict_output", conf_threshold_var.get(), task_controller=task_controller)
                print("Detect and Crop completed!")

            # Chạy Read and Export
            if task_controller.is_running():
                print("Running Export to Excel...")
                read_text_from_images("./predict_output", output_excel_var.get(), task_controller=task_controller)
                print("Read and Export completed!")

        run_task_in_thread(task, task_controller)

    def stop_task():
        """Dừng tiến trình hiện tại."""
        task_controller.stop()
        print("Task stopped.")

    # GUI Layout
    tk.Label(root, text="Input Folder:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=input_folder_var, width=50).grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=select_input_folder).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="YOLO Model File:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=model_path_var, width=50).grid(row=1, column=1, padx=10, pady=5)
    tk.Button(root, text="Browse", command=select_model_file).grid(row=1, column=2, padx=10, pady=5)

    tk.Label(root, text="Output Excel File:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=output_excel_var, width=50).grid(row=2, column=1, padx=10, pady=5)

    tk.Label(root, text="Confidence Threshold:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=conf_threshold_var, width=10).grid(row=3, column=1, padx=10, pady=5, sticky="w")

    # Adjustable parameters
    tk.Label(root, text="Lower HSV1:").grid(row=4, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=lower_hsv1_var, width=20).grid(row=4, column=1, padx=10, pady=5, sticky="w")

    tk.Label(root, text="Upper HSV1:").grid(row=5, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=upper_hsv1_var, width=20).grid(row=5, column=1, padx=10, pady=5, sticky="w")

    tk.Label(root, text="Lower HSV2:").grid(row=6, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=lower_hsv2_var, width=20).grid(row=6, column=1, padx=10, pady=5, sticky="w")

    tk.Label(root, text="Upper HSV2:").grid(row=7, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=upper_hsv2_var, width=20).grid(row=7, column=1, padx=10, pady=5, sticky="w")

    tk.Label(root, text="Kernel Size:").grid(row=8, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=kernel_size_var, width=20).grid(row=8, column=1, padx=10, pady=5, sticky="w")

    tk.Label(root, text="Threshold:").grid(row=9, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=threshold_var, width=10).grid(row=9, column=1, padx=10, pady=5, sticky="w")

    tk.Button(root, text="Convert to BW", command=run_convert_to_bw, width=20, height=2).grid(row=10, column=0, pady=10)
    tk.Button(root, text="Detect and Crop", command=run_detect_and_crop, width=20, height=2).grid(row=10, column=1, pady=10)
    tk.Button(root, text="Read and Export", command=run_read_and_export, width=20, height=2).grid(row=10, column=2, pady=10)
    tk.Button(root, text="Run All", command=run_all_tasks, width=20, height=2).grid(row=11, column=0, columnspan=3, pady=10)
    tk.Button(root, text="Stop", command=stop_task, width=20, height=2).grid(row=12, column=0, columnspan=3, pady=20)

    # Add a TextBox to display console output
    console_output = ScrolledText(root, wrap=tk.WORD, height=15, width=80)
    console_output.grid(row=13, column=0, columnspan=3, padx=10, pady=10)

    # Redirect console output to the TextBox
    sys.stdout = ConsoleRedirector(console_output)
    sys.stderr = ConsoleRedirector(console_output)

    root.mainloop()

if __name__ == "__main__":
    run_app()